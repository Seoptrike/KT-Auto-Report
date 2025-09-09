from __future__ import annotations
from io import BytesIO
from pathlib import Path
from typing import Dict, Tuple, Optional
from datetime import datetime, date
import re
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from calendar import monthrange
import math
from collections import defaultdict

# (옵션) Flask 응답이 필요할 때만 쓰세요
try:
    from flask import send_file as _flask_send_file  # type: ignore
except Exception:
    _flask_send_file = None


# ===== 공통 유틸 =====
def _make_yyyy_mm_dir(base_dir: Path | str, when: Optional[datetime] = None) -> Path:
    when = when or datetime.now()
    base = Path(base_dir)
    out_dir = base / when.strftime("%Y") / when.strftime("%m")
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _parse_yyyymm(val) -> Optional[datetime]:
    """
    A2 셀 값에서 YYYY-MM(또는 YYYY.MM / YYYY/MM / 'YYYYMM' 등) 파싱 → 해당 월의 1일 datetime 반환.
    날짜/문자열 모두 대응.
    """
    if isinstance(val, datetime):
        return datetime(val.year, val.month, 1)
    if isinstance(val, date):
        return datetime(val.year, val.month, 1)
    if isinstance(val, (int, float)):
        s = str(int(val))
    else:
        s = str(val or "").strip()

    if not s:
        return None

    # 1) YYYY-MM 계열
    m = re.match(r"^\s*(\d{4})[.\-\/ ]?(\d{1,2})\s*$", s)
    if m:
        y, mm = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12:
            return datetime(y, mm, 1)

    # 2) YYYYMM 붙어있는 형태
    m = re.match(r"^\s*(\d{4})(\d{2})\s*$", s)
    if m:
        y, mm = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12:
            return datetime(y, mm, 1)

    return None

def _read_a2_month(file_like) -> Optional[datetime]:
    """
    원본 첫 시트의 A2 값을 읽어 base month(해당월 1일)로 반환.
    file_like는 BytesIO/스트림 모두 가능.
    """
    wb = load_workbook(file_like, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    a2 = ws["A2"].value
    return _parse_yyyymm(a2)

def split_vat_exact(gross_list):
    raw = [g * 10 / 11 for g in gross_list]
    floor_vals = [math.floor(x) for x in raw]
    residues = [x - f for x, f in zip(raw, floor_vals)]

    target = round(sum(raw))           # 정책에 따라 floor/ceil/half_up로 변경 가능
    need = target - sum(floor_vals)

    order = sorted(range(len(residues)), key=lambda i: residues[i], reverse=True)
    for i in order[:need]:
        floor_vals[i] += 1

    supply = [int(v) for v in floor_vals]                    # ← 정수화
    vat = [int(round(g)) - s for g, s in zip(gross_list, supply)]  # ← 정수로 계산
    return supply, vat

# ===== 1) 원본에서 추출 =====
def extract_p_bi_mapped_only(
    file_like,
    name_col: str = "P",
    amount_col: str = "BI",
) -> Dict[str, int]:
    raw = file_like.read() if hasattr(file_like, "read") else file_like
    bio = BytesIO(raw if isinstance(raw, (bytes, bytearray)) else raw.getvalue())
    wb = load_workbook(bio, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    patterns = [
        (re.compile(r"남"), "㈜남이섬"),
        (re.compile(r"엠지브이보"), "㈜엠지브이보안시스템"),
        (re.compile(r"즐거"), "㈜즐거운세상"),
        (re.compile(r"더"), "유한회사 더늘푸른"),
    ]
    def canonize(name: str) -> Optional[str]:
        if not name: return None
        for pat, canon in patterns:
            if pat.search(name): return canon
        return None

    # 1) 행 순서대로 (업체, 총액) 모으기
    rows: list[tuple[str, float]] = []
    for r in range(2, ws.max_row + 1):
        nval = ws[f"{name_col}{r}"].value
        aval = ws[f"{amount_col}{r}"].value
        name = str(nval).strip() if nval not in (None, "") else ""
        canon = canonize(name)
        if not canon: continue
        if aval in (None, ""): continue
        try:
            amt = float(str(aval).replace(",", ""))
        except Exception:
            continue
        if amt > 0:
            rows.append((canon, amt))

    if not rows:
        return {}

    # 2) ★ 전체 한 번에 정확 배분 → 공급가 리스트
    gross_all = [amt for _, amt in rows]
    supply_all, _ = split_vat_exact(gross_all)  # ← 전체 합이 304,545로 맞춰짐

    # 3) 업체별 공급가 합계로 집계
    result: Dict[str, int] = defaultdict(int)
    for (vendor, _), supply in zip(rows, supply_all):
        result[vendor] += int(supply)

    return dict(result)


# ===== 2) 템플릿 채워 메모리로 만들기 =====
def build_sample2_bytes(
    mapped_result: Dict[str, float],
    template_path: str | Path,
    write_formulas: bool = False,
    filename_date_fmt: str = "dots",   # 무시 가능, 파일명은 아래서 YY.MM.dd 고정
    settlement_month: Optional[datetime] = None,
    report_date: Optional[datetime] = None,
) -> Tuple[BytesIO, str]:
    wb = load_workbook(str(template_path))
    ws = wb.active

    # 정산월(없으면 -3개월 fallback 대신 지금 시점에서 -3개월은 이전 로직이었음)
    target_date = settlement_month or (datetime.now() - relativedelta(months=3))

    # 표시용 날짜: 보고일자 우선, 없으면 정산월 사용
    use_date = report_date or target_date

    # 한국어 요일 헬퍼
    weekdays_kr = ["월", "화", "수", "목", "금", "토", "일"]
    def fmt_kor(dt: datetime) -> str:
        return f"{dt.year}년 {dt.month}월 {dt.day}일 {weekdays_kr[dt.weekday()]}요일"

    # ★ D8 = YYYY년 M월 D일 요일 (앞자리 0 제거)
    ws["D8"] = fmt_kor(use_date)

    # ★ D6 = YY년 (정산월-3개월)월 A'cen Cloud 판매위탁 협력사 정산 (3개월전)
    adjusted_date = target_date - relativedelta(months=3)
    yy2 = adjusted_date.strftime("%y")
    mm2 = adjusted_date.month  # int → 0 없음
    ws["D6"] = f"{yy2}년 {mm2}월 A'cen Cloud 판매위탁 협력사 정산"

    # ★ G7 = (D8 기준) 다음달의 마지막 날을 같은 포맷으로
    next_month = use_date + relativedelta(months=1)
    from calendar import monthrange
    last_day = monthrange(next_month.year, next_month.month)[1]
    end_of_next_month = next_month.replace(day=last_day)
    ws["G7"] = fmt_kor(end_of_next_month)

    # 본문 라인: (예시) 정산 문구는 정산월-3개월 기준으로 유지
    yy_for_body = adjusted_date.strftime("%y")
    mm_for_body = adjusted_date.month  # int (9월 → 9)
    start_row = 11
    seq = 1
    for i, (name, amount) in enumerate(mapped_result.items(), start=start_row):
        ws[f"A{i}"] = seq
        ws[f"D{i}"] = f"A'CenCloud {name} 정산({yy_for_body}년 {mm_for_body}월) - 판매수수료"
        ws[f"C{i}"] = "판매위탁 수수료 (A'cen)"
        ws[f"E{i}"] = 1
        ws[f"K{i}"] = float(amount)
        ws[f"G{i}"] = "-"
        seq += 1

    # 저장용 바이트 + 제안 파일명(YY.MM.dd)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    date_label = (use_date).strftime("%y.%m.%d")
    suggested = f"매출결의서_KT ACen_유지수수료_{date_label}.xlsx"
    return bio, suggested

# ===== 3) YYYY/MM 폴더에 저장 =====
def save_acen_bytes_yyyy_mm(
    bio: BytesIO,
    filename_base: str = "매출결의서_KT ACen_유지수수료",
    base_dir: Path | str = "output",
    when: Optional[datetime] = None,            # 폴더(YYYY/MM) 기준: 정산월
    date_fmt: str = "dots",                     # "dots" | "underscores"
    report_date: Optional[datetime] = None,     # 파일명 날짜 기준(우선)
) -> Path:
    # 1) 디렉토리: 정산월(when) 기준
    dir_basis = when or datetime.now()
    out_dir = _make_yyyy_mm_dir(base_dir, dir_basis)

    # 2) 파일명: report_date 우선, 없으면 정산월(when)
    name_basis = report_date or dir_basis
    date_str = (
        name_basis.strftime("%y_%m_%d")
        if date_fmt == "underscores"
        else name_basis.strftime("%y.%m.%d")
    )

    out_path = out_dir / f"{filename_base}_{date_str}.xlsx"
    with open(out_path, "wb") as f:
        f.write(bio.getbuffer())
    return out_path

def run_acen_pipeline(
    file_like,
    template_path: str | Path,
    base_dir: Path | str = "output",
    write_formulas: bool = False,
    date_fmt: str = "dots",
    filename_base: str = "매출결의서_KT ACen_유지수수료",
    when: Optional[datetime] = None,           # 그대로 두되 사용 안 함
    report_day: Optional[int] = None,          # ★ 추가: 일(day)만 받기
) -> Path:
    # 업로드 원본 복사
    raw = file_like.read() if hasattr(file_like, "read") else file_like
    buf1, buf2 = BytesIO(raw), BytesIO(raw)

    # 1) 데이터 집계
    mapped = extract_p_bi_mapped_only(buf1)

    # 2) 정산월 계산 (A2 + 1개월, 실패 시 now-3M)
    base_month = _read_a2_month(buf2)
    settlement_month = base_month + relativedelta(months=1) if base_month else (datetime.now() - relativedelta(months=3))

    # 3) report_day → report_date 생성 (해당 월에 없는 일자면 말일로 보정)
    report_date = None
    if isinstance(report_day, int) and report_day > 0:
        y, m = settlement_month.year, settlement_month.month
        last_day = monthrange(y, m)[1]
        safe_day = min(report_day, last_day)
        report_date = settlement_month.replace(day=safe_day)

    # 4) 템플릿 채우기
    bio, _ = build_sample2_bytes(
        mapped_result=mapped,
        template_path=template_path,
        write_formulas=write_formulas,
        filename_date_fmt=("underscores" if date_fmt == "underscores" else "dots"),
        settlement_month=settlement_month,
        report_date=report_date,  # ★ 전달!
    )

    # 5) 저장 (정산월 기준 경로/파일명)
    return save_acen_bytes_yyyy_mm(
        bio=bio,
        filename_base=filename_base,
        base_dir=base_dir,
        when=settlement_month,
        date_fmt=date_fmt,
        report_date=report_date, 
    )