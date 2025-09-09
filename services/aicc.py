# services/aicc.py
from __future__ import annotations
from pathlib import Path
from collections import OrderedDict
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, List, Sequence, Tuple, Optional
import fnmatch
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date
from io import BytesIO
import re
from calendar import monthrange
from dateutil.relativedelta import relativedelta

# =========================
# 유틸
# =========================
def _norm(x):
    if pd.isna(x):
        return None
    if isinstance(x, str):
        s = x.strip()
        return None if s == "" else s
    return x

def _to_decimal(x) -> Decimal:
    if x is None: return Decimal("0")
    if isinstance(x, Decimal): return x
    if isinstance(x, (int, float)): return Decimal(str(x))
    if isinstance(x, str):
        s = x.replace(",", "").strip()
        return Decimal(s) if s else Decimal("0")
    return Decimal("0")

def _parse_yyyymm(val) -> Optional[datetime]:
    """
    A2 셀 값에서 YYYY-MM(또는 YYYY.MM / YYYY/MM / 'YYYYMM' 등) 파싱 → 해당 월의 1일 datetime 반환.
    날짜/문자열 모두 대응.
    """
    if isinstance(val, datetime):
        return datetime(val.year, val.month, 1)
    if isinstance(val, date):
        return datetime(val.year, val.month, 1)
    if isinstance(val, (int, float)):
        s = str(int(val))
    else:
        s = str(val or "").strip()

    if not s:
        return None

    # 1) YYYY-MM 계열
    m = re.match(r"^\s*(\d{4})[.\-\/ ]?(\d{1,2})\s*$", s)
    if m:
        y, mm = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12:
            return datetime(y, mm, 1)

    # 2) YYYYMM 붙어있는 형태
    m = re.match(r"^\s*(\d{4})(\d{2})\s*$", s)
    if m:
        y, mm = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12:
            return datetime(y, mm, 1)

    return None

# =========================
# Read & Combine
# =========================
def read_bghm_one(file_obj, start_row: int = 7) -> Tuple[List[List[Any]], Optional[datetime]]:
    """
    단일 파일에서 B/G/H/M을 start_row부터 읽어서 [[B,G,H,M], ...] 리턴
    + 첫 시트 A4의 정산년월(YYYYMM 등)을 파싱해 해당 월의 1일 datetime도 함께 리턴
    """
    # 1) 파일을 메모리에 복사해 두 개의 버퍼로 사용 (pandas / openpyxl 각각)
    raw = file_obj.read() if hasattr(file_obj, "read") else file_obj
    if isinstance(raw, (bytes, bytearray)):
        raw_bytes = raw
    else:
        # raw가 BytesIO인 경우 등
        raw_bytes = raw.getvalue()
    bio_pd = BytesIO(raw_bytes)   # pandas용
    bio_xl = BytesIO(raw_bytes)   # openpyxl용

    # 2) pandas로 본문(B,G,H,M) 읽기
    df = pd.read_excel(
        bio_pd,
        sheet_name=0,
        usecols="B,G,H,M",
        header=None,
        skiprows=start_row - 1,
        engine="openpyxl",
        dtype=object,
    )
    df.columns = ["B", "G", "H", "M"]
    try:
        df = df.map(_norm)              # pandas>=2.2
    except Exception:
        df = df.applymap(_norm)         # fallback

    df = df.dropna(how="all")
    df = df.dropna(subset=["B", "G", "H", "M"])
    rows = df.values.tolist()

    # 3) openpyxl로 A4 읽어서 정산월 파싱
    wb = load_workbook(bio_xl, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    a4 = ws["A4"].value

    # _parse_yyyymm은 기존에 있으니 재사용 (YYYYMM/YYYY-MM/… 대응, 해당 월 1일 반환)
    settlement_month = _parse_yyyymm(a4)  # Optional[datetime]

    return rows, settlement_month

def combine_bghm_from_paths(
    paths_or_files: Sequence[Any],
    start_row: int = 7,
    strict_same_month: bool = False,
) -> Tuple[List[List[Any]], Optional[datetime]]:
    """
    여러 파일에서 [[B,G,H,M], ...] 병합 + 정산월 반환
    - 각 파일의 첫 시트 A4에서 정산월(YYYYMM 등) 파싱 → 해당 월 1일(datetime)
    - 모든 파일의 정산월이 다르면:
        * strict_same_month=True: ValueError
        * strict_same_month=False: 가장 최신 월을 선택하고 경고 출력
    반환: (combined_rows, chosen_settlement_month)
    """
    combined: List[List[Any]] = []
    months: List[datetime] = []

    for item in paths_or_files:
        if hasattr(item, "read"):  # file-like
            rows, month = read_bghm_one(item, start_row=start_row)  # ← f 아님!
        else:
            p = Path(item)
            if not p.exists():
                print(f"[WARN] 파일 없음: {p}")
                continue
            with open(p, "rb") as f:
                rows, month = read_bghm_one(f, start_row=start_row)

        if rows:
            combined.extend(rows)
        if month:
            months.append(month)

    chosen: Optional[datetime] = None
    if months:
        # year, month 쌍 기준으로 상이 여부 판단
        ym_set = {(m.year, m.month) for m in months}
        if len(ym_set) > 1:
            msg = "[WARN] A4 정산월이 파일마다 다릅니다: " + ", ".join(
                sorted({f"{y}-{m:02d}" for y, m in ym_set})
            )
            if strict_same_month:
                raise ValueError(msg)
            else:
                print(msg + " → 가장 최신 월로 선택합니다.")
        # 가장 최신(큰) 월 선택
        chosen = max(months, key=lambda d: (d.year, d.month))

    return combined, chosen

# =========================
# Type/Title 생성
# =========================
def infer_type_from_m(m) -> str | None:
    if m is None:
        return None
    s = str(m)
    su = s.upper()
    if "IB" in su:
        return "IB"
    if "OB" in su:
        return "OB"
    if "챗봇" in s:
        return "챗봇"
    return None

def build_title(g, h, typ: str | None) -> str:
    """
    규칙:
    1) H == 3*G → 모집수수료
    2) H != 3*G && H > 1,000,000 → 개발비용
    3) H != 3*G && H < 1,000,000 && OB → 회선수수료
    4) H != 3*G && H < 1,000,000 && IB → 유지수수료
    5) H != 3*G && H < 1,000,000 && 챗봇 → {Type} 유지수수료
    """
    G = _to_decimal(g)
    H = _to_decimal(h)
    EPS = Decimal("0.000001")
    type_disp = typ if typ else "미정"

    if (H - (G * 3)).copy_abs() <= EPS:
        return f"판매위탁 수수료 (A'cen) 보이스봇({type_disp}) 모집수수료"

    if H > Decimal("1000000"):
        return f"판매위탁 수수료 (A'cen) 보이스봇({type_disp}) 개발비용"

    if (typ or "").upper() == "OB":
        return f"판매위탁 수수료 (A'cen) 보이스봇({type_disp}) 회선수수료"
    if (typ or "").upper() == "IB":
        return f"판매위탁 수수료 (A'cen) 보이스봇({type_disp}) 유지수수료"
    if typ == "챗봇":
        return f"판매위탁 수수료 (A'cen) {type_disp} 유지수수료"

    return f"판매위탁 수수료 (A'cen) 보이스봇({type_disp}) 기타"

def enrich_bghm_rows(rows: Sequence[Sequence[Any]]) -> List[List[Any]]:
    """
    [[B,G,H,M]] → [[B,G,H,M,Type,Title]]
    """
    out: List[List[Any]] = []
    for r in rows:
        if not r or len(r) < 4:
            continue
        B, G, H, M = r[0], r[1], r[2], r[3]
        typ = infer_type_from_m(M)
        title = build_title(G, H, typ)
        out.append([B, G, H, M, typ, title])
    return out

# =========================
# 그룹핑 & 매핑
# =========================
def group_sum_by_name_title_H(rows: Sequence[Sequence[Any]], keep_order: bool = True) -> List[List[Any]]:
    """
    [[B,G,H,M,Type,Title]] → (B,Title) 기준 H 합산
    반환: [[B, total_H, Title]]
    """
    acc: dict[Tuple[Any, Any], Decimal] = OrderedDict() if keep_order else {}
    for r in rows:
        if not r or len(r) < 6:
            continue
        name, H, title = r[0], r[2], r[5]
        if name is None or title is None:
            continue
        key = (name, title)
        acc[key] = acc.get(key, Decimal("0")) + _to_decimal(H)

    out: List[List[Any]] = []
    for (name, title), total in acc.items():
        rounded = total.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        out.append([name, int(rounded), title])
    return out

NAME_MAP_RULES = [
    ("(주)오토피*", "오토피온"),
    ("(주)캐럿솔루션*", "캐럿솔루션즈"),
    ("(주)이앤에이치 에너*", "이앤에이치에너지"),
    ("주식회사 브리지*", "브리지텍"),
    ("(주)엠지브이보안시스*", "엠지브이보안시스템"),
    ("순천향대학교부속서울병*", "순천향대병원"),
    ("충남신용보증재*", "충남신용보증재단"),
    ("전북신용보증재*", "전북신용보증재단"),
    ("주식회사 메타전*", "메타전스"),
    ("대전서구*", "대전서구청"),
    ("익산시*", "익산시청"),
    ("유성구*", "유성구청"),
    ("웰스라이*", "웰스라이프"),
]

def apply_name_mapping(name: str) -> str:
    if not name:
        return name
    for pattern, replacement in NAME_MAP_RULES:
        if fnmatch.fnmatch(name, pattern):
            return replacement
    return name

def map_grouped_names(grouped: Sequence[Sequence[Any]]) -> List[List[Any]]:
    out: List[List[Any]] = []
    for r in grouped:
        if not r or len(r) < 3:
            continue
        out.append([apply_name_mapping(r[0]), r[1], r[2]])
    return out

# =========================
# 엑셀 쓰기
# =========================
def write_to_excel(
    mapped: dict,
    template_path: str | Path,
    base_dir: str | Path,
    *,
    settlement_month: datetime | None = None,
    report_day: int | None = None,
) -> Path:
    """
    mapped_grouped: [[회사명, total_H, Title], ...]
    템플릿의 A/C/D/E/K/G 열에 기록 후 "매출결의서_KT AICC_YY.mm.dd.xlsx"로 저장
    """
    # 1) 정산월 기본값
    target = settlement_month or datetime.now()

    # 2) report_date 만들기 (정산월 내에서 day만 교체, 말일 보정)
    report_date = None
    if isinstance(report_day, int) and report_day > 0:
        last = monthrange(target.year, target.month)[1]
        report_date = target.replace(day=min(report_day, last))

    wb = load_workbook(template_path)
    ws = wb.active

     # 요일 포맷
    weekdays_kr = ["월", "화", "수", "목", "금", "토", "일"]
    def fmt_kor(dt: datetime) -> str:
        return f"{dt.year}년 {dt.month}월 {dt.day}일 {weekdays_kr[dt.weekday()]}요일"

    use_date = report_date or target
    ws["D8"] = fmt_kor(use_date)  # YYYY년 M월 D일 요일

    nm = use_date + relativedelta(months=1)
    from calendar import monthrange as _mr
    eonm = nm.replace(day=_mr(nm.year, nm.month)[1])
    ws["G7"] = fmt_kor(eonm)

    yy = (report_date or settlement_month or datetime.now()).strftime("%y")
    ws["D6"] = f"{yy}년 A'CenCloud 사용량 판매위탁 협력사 정산"

    start_row = 11
    seq = 1

    for i, row in enumerate(mapped, start=start_row):
        name, total_h, title = row[0], row[1], row[2]
        ws[f"A{i}"] = seq
        ws[f"C{i}"] = title
        ws[f"D{i}"] = f"A'CenCloud 사용량 판매위탁 협력사 정산({name})"
        ws[f"E{i}"] = 1
        ws[f"K{i}"] = total_h
        ws[f"G{i}"] = "-"  # 필요시 제거 가능
        seq += 1

     # === 저장 디렉토리: base_dir/YYYY/MM ===
    if base_dir is None:
        base_dir = template_path.parent
    year_dir = base_dir / use_date.strftime("%Y") / use_date.strftime("%m")
    year_dir.mkdir(parents=True, exist_ok=True)

    today_str = use_date.strftime("%y.%m.%d")
    out_name = f"매출결의서_KT AICC_{today_str}.xlsx"
    out_path = year_dir / out_name

    wb.save(out_path)
    return out_path
