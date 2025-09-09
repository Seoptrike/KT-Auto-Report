# sum.py
from pathlib import Path
from openpyxl import load_workbook
import re
from typing import Optional
from collections import defaultdict
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet
from calendar import monthrange
from dateutil.relativedelta import relativedelta

OUTPUT_DIR = Path("output")
AICC_PATTERN = re.compile(r"매출결의서_KT AICC_.*\.xlsx$")
ACEN_PATTERN = re.compile(r"매출결의서_KT ACen_유지수수료_.*\.xlsx$")

def _make_yyyy_mm_dir(base_dir: Path | str, when: Optional[datetime] = None) -> Path:
    when = when or datetime.now()
    base = Path(base_dir)
    out_dir = base / when.strftime("%Y") / when.strftime("%m")
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def find_latest_file_for_month(
    base_dir: str | Path,
    when: datetime,
    prefix: str,                    # 예: "매출결의서_KT ACen" / "매출결의서_KT AICC"
    ext: str = ".xlsx",
) -> Optional[Path]:
    month_dir = Path(base_dir) / when.strftime("%Y") / when.strftime("%m")
    if not month_dir.exists():
        return None
    # 접두어(prefix)로 필터링
    candidates = [p for p in month_dir.glob(f"{prefix}*{ext}") if p.is_file()]
    if not candidates:
        return None
    # 수정시간 최신(또는 파일명 날짜 파싱으로 정렬해도 OK)
    return max(candidates, key=lambda p: p.stat().st_mtime)

def extract_D_K_rows(xlsx_path: Path, start_row: int = 11, end_row: int = 25):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    out = []
    for r in range(start_row, end_row + 1):
        d = ws[f"D{r}"].value
        k = ws[f"K{r}"].value
        if (d is None or (isinstance(d, str) and not d.strip())) and \
           (k is None or (isinstance(k, str) and not str(k).strip())):
            continue
        out.append((d, k))
    return out


def pretty_print_pairs(pairs, title: str):
    print(f"\n=== {title} (rows={len(pairs)}) ===")
    for i, (d, k) in enumerate(pairs, start=1):
        print(f"  {i:02d}. D: {d}    |    K: {k}")

def clean_company_name(name: str | None) -> str | None:
    if not name:
        return None
    # 제거할 패턴들
    patterns = [r"^\(주\)\s*", r"^㈜\s*", r"^유한회사\s*"]
    cleaned = name
    for p in patterns:
        cleaned = re.sub(p, "", cleaned)
    return cleaned.strip()

def extract_company_aicc(d_value: str) -> str | None:
    """AICC: 괄호 안의 회사명 추출"""
    if not d_value:
        return None
    m = re.search(r"\((.*?)\)", d_value)
    if not m:
        return None
    return clean_company_name(m.group(1))

def extract_company_acen(d_value: str) -> str | None:
    """ACEN: A'CenCloud ~ 정산( 직전의 회사명 추출"""
    if not d_value:
        return None
    m = re.search(r"A'CenCloud\s+(.*?)\s+정산\(", d_value)
    if not m:
        return None
    return clean_company_name(m.group(1))

def merge_by_company(rows: list[tuple[str, float]]) -> list[tuple[str, float]]:
    """
    rows: [(회사명, 금액), ...]
    같은 회사명은 금액 합산.
    반환: [(회사명, 총금액), ...]
    """
    acc = defaultdict(float)
    for name, amount in rows:
        if not name:
            continue
        acc[name] += float(amount or 0)

    # 정렬은 원하면 회사명 알파벳순/금액순 등 지정 가능
    return list(acc.items())

NAME_MAP_SUM = {
    "캐럿솔루션즈": "당근영어(캐럿솔루션즈)",
    "이앤에이치에너지": "E&H에너지",
    "브리지텍": "중소기업중앙회(브리지텍)",
    "엠지브이보안시스템": "MGV보안시스템",
    "메타전스": "대전서구청_스마트통신팀(메타전스)",
    "대전서구청": "대전서구청_주차행정과",
    "즐거운세상": "인터불고호텔",
}

def apply_sum_name_mapping(rows: list[tuple[str, float]]) -> list[tuple[str, float]]:
    """
    rows: [(회사명, 금액), ...]
    NAME_MAP_SUM 기준으로 이름 치환. 없으면 그대로 유지.
    """
    out = []
    for name, amount in rows:
        mapped = NAME_MAP_SUM.get(name, name)  # 매핑 있으면 바꿔주고, 없으면 그대로
        out.append((mapped, amount))
    return out

def _parse_month_cell(v) -> int | None:
    """
    L4:W4 셀 값이 숫자(1~12) 혹은 '1월', '01', '01월' 등일 수 있으니
    숫자만 뽑아서 int로 변환.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        iv = int(v)
        return iv if 1 <= iv <= 12 else None
    s = str(v).strip()
    m = re.search(r"(\d{1,2})", s)
    if not m:
        return None
    iv = int(m.group(1))
    return iv if 1 <= iv <= 12 else None

def _build_name_row_map(ws, start_row=5, end_row=24, name_col="B") -> dict[str, int]:
    """
    B5:B24의 회사명 → 엑셀 행번호 매핑 딕셔너리 생성.
    공백 트림해서 매칭 안정화.
    """
    name_to_row = {}
    for r in range(start_row, end_row + 1):
        v = ws[f"{name_col}{r}"].value
        if v is None:
            continue
        key = str(v).strip()
        if key:
            name_to_row[key] = r
    return name_to_row

def _clear_year_data(ws: Worksheet, start_row=5, end_row=24,
                     start_col_letter="L", end_col_letter="W"):
    """
    월별 값 영역(L5:W24)을 모두 비움(None).
    회사명(B5:B24)과 헤더(L4:W4)는 유지.
    """
    for r in range(start_row, end_row + 1):
        for c in range(ord(start_col_letter), ord(end_col_letter) + 1):
            ws[f"{chr(c)}{r}"].value = None

def fill_sum_template(
    mapped: dict,
    template_path: str | Path,
    out_base_dir: str | Path,
    *,
    settlement_month: datetime | None = None,   # ★ 전달된 정산월(전달)
    report_day: int | None = None,              # ★ 일(day)만
    date_fmt: str = "dots",
) -> Path:
    # 1) 기준월 = 정산월(없으면 now)
    target = settlement_month or datetime.now()

    # 2) report_date (정산월 내에서 day만 교체, 말일 보정)
    report_date = None
    if isinstance(report_day, int) and report_day > 0:
        last = monthrange(target.year, target.month)[1]
        report_date = target.replace(day=min(report_day, last))

    # === (NEW) 전달 파일을 우선 템플릿으로 사용 ===
    prev_month = target - relativedelta(months=1)
    prev_dir = Path(out_base_dir) / prev_month.strftime("%Y") / prev_month.strftime("%m")
    prev_date_str = (
        prev_month.strftime("%y_%m") if date_fmt == "underscores" else prev_month.strftime("%y.%m")
    )
    prev_candidate = prev_dir / f"업무실적_{prev_date_str}.xlsx"

    try:
        if prev_candidate.exists():
            wb = load_workbook(prev_candidate)
        else:
            wb = load_workbook(template_path)
    except Exception:
        # 혹시라도 열기 실패하면 기본 템플릿으로 폴백
        wb = load_workbook(template_path)

    ws = wb.active

    # ---- A2 제목: 정산년도 기준으로 관리 ----
    yyyy = target.strftime("%Y")
    month_num = target.month  # 1~12

    if month_num == 1:
        _clear_year_data(ws, start_row=5, end_row=24, start_col_letter="L", end_col_letter="W")
        ws["A2"] = f"{yyyy}년 KT AICC 실적 현황"
    else:
        if not ws["A2"].value:
            ws["A2"] = f"{yyyy}년 KT AICC 실적 현황"

    # ---- L4:W4에서 '정산월'에 해당하는 열 찾기 ----
    header_cells = [f"{col}4" for col in [chr(c) for c in range(ord('L'), ord('W')+1)]]
    target_col_letter = None
    for addr in header_cells:
        iv = _parse_month_cell(ws[addr].value)
        if iv == month_num:
            target_col_letter = addr[0]
            break
    if target_col_letter is None:
        raise RuntimeError(f"L4:W4에서 {month_num}월에 해당하는 열을 찾지 못했습니다.")

    # ---- B5:B24 회사명 → 행번호 맵 ----
    name_to_row = _build_name_row_map(ws, start_row=5, end_row=24, name_col="B")

    # ---- mapped 값 쓰기 ----
    items = mapped.items() if isinstance(mapped, dict) else mapped
    missing = []
    for name, amount in items:
        key = str(name).strip()
        r = name_to_row.get(key)
        if r is None:
            missing.append(key)
            continue
        ws[f"{target_col_letter}{r}"] = float(amount)

    # ---- 저장: 폴더=정산월(YYYY/MM), 파일명=보고일(없으면 정산월) ----
    out_dir = _make_yyyy_mm_dir(Path(out_base_dir), target)
    basis = report_date or target
    date_str = basis.strftime("%y_%m") if date_fmt == "underscores" else basis.strftime("%y.%m")
    out_name = f"업무실적_{date_str}.xlsx"
    out_path = out_dir / out_name

    wb.save(out_path)

    if missing:
        print("[WARN] 템플릿 B5:B24에서 못 찾은 이름들:", ", ".join(sorted(set(missing))))

    return out_path
